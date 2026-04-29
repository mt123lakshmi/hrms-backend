from fastapi import HTTPException
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select
from sqlalchemy.orm import selectinload
from datetime import datetime
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
import io

from app.models.employee_asset import EmployeeAsset
from app.models.employee import Employee   # 🔥 ADDED


# ===============================
# 🔹 GET ASSETS (TABLE)
# ===============================
async def get_assets_controller(db: AsyncSession, user):   # 🔥 ADDED user

    result = await db.execute(
        select(EmployeeAsset)
        .join(Employee, EmployeeAsset.employee_id == Employee.id)   # 🔥 ADDED
        .options(selectinload(EmployeeAsset.employee))
        .where(Employee.company_id == user.company_id)   # 🔥 ADDED
    )

    assets = result.scalars().all()

    data = []

    for a in assets:
        data.append({
            "asset_id": a.id,
            "employee_id": a.employee.id if a.employee else None,
            "employee_name": a.employee.name if a.employee else None,
            "laptop_asset_id": a.laptop_asset_id,
            "access_card": a.access_card,
            "additional_asset": a.additional_asset,
            "status": a.status,
            "unassign_reason": a.return_reason if a.status == "unassigned" else None
        })

    return {"success": True, "data": data}


# ===============================
# 🔹 ASSIGN / UNASSIGN
# ===============================
async def asset_action_controller(asset_id: int, data, db: AsyncSession, user):  # 🔥 ADDED user

    result = await db.execute(
        select(EmployeeAsset)
        .join(Employee, EmployeeAsset.employee_id == Employee.id)   # 🔥 ADDED
        .where(
            EmployeeAsset.id == asset_id,
            Employee.company_id == user.company_id   # 🔥 ADDED
        )
    )
    asset = result.scalar_one_or_none()

    if not asset:
        raise HTTPException(status_code=404, detail="Asset not found")

    # ===============================
    # ASSIGN
    # ===============================
    if data.action == "assign":

        if asset.status == "assigned":
            return {
                "success": False,
                "message": "Asset already assigned"
            }

        asset.status = "assigned"
        asset.assigned_reason = data.reason
        asset.assigned_at = datetime.utcnow()

    # ===============================
    # UNASSIGN
    # ===============================
    elif data.action == "unassign":

        if asset.status == "unassigned":
            return {
                "success": False,
                "message": "Asset already unassigned"
            }

        asset.status = "unassigned"
        asset.return_reason = data.reason
        asset.returned_at = datetime.utcnow()

    await db.commit()

    return {
        "success": True,
        "message": f"Asset {data.action} successful"
    }


# ===============================
# 🔹 DOWNLOAD EXCEL
# ===============================
async def download_assets_excel_controller(db: AsyncSession, user):  # 🔥 ADDED user

    result = await db.execute(
        select(EmployeeAsset)
        .join(Employee, EmployeeAsset.employee_id == Employee.id)   # 🔥 ADDED
        .options(selectinload(EmployeeAsset.employee))
        .where(Employee.company_id == user.company_id)   # 🔥 ADDED
    )

    assets = result.scalars().all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Assets"

    # =========================
    # HEADER
    # =========================
    headers = [
        "Asset ID",
        "Employee",
        "Laptop ID",
        "Access Card",
        "Additional Assets",
        "Status",
        "Unassign Reason"
    ]

    ws.append(headers)

    # Bold header
    for col in range(1, len(headers) + 1):
        ws.cell(row=1, column=col).font = Font(bold=True)

    # =========================
    # DATA
    # =========================
    for a in assets:
        raw_assets = a.additional_asset or ""

        asset_list = [x.strip() for x in raw_assets.split(",") if x.strip()]
        formatted_assets = "\n".join(asset_list)

        ws.append([
            a.id,
            a.employee.name if a.employee else "",
            a.laptop_asset_id,
            a.access_card,
            formatted_assets,
            a.status,
            a.return_reason if a.status == "unassigned" else ""
        ])

        cell = ws.cell(row=ws.max_row, column=5)
        cell.alignment = Alignment(wrap_text=True)

    # =========================
    # AUTO WIDTH
    # =========================
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter

        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))

        ws.column_dimensions[col_letter].width = min(max_length + 2, 50)

    # =========================
    # STREAM FILE
    # =========================
    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)

    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": "attachment; filename=employee_assets.xlsx"
        }
    )